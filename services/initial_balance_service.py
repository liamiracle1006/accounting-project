"""
AgentLedger V4.0 — InitialBalanceService (Sprint 2.2)

职责：
  1. save_balance()                    — 保存单条期初余额（含 year_start 推导、聚合）
  2. batch_save()                      — 批量保存（整体事务，一行失败全部回滚）
  3. get_balances_with_subjects()      — 科目树 + 期初余额联合查询
  4. calculate_trial_balance()         — 四维度本位币试算平衡
  5. calculate_foreign_trial_balance() — 外币独立试算平衡
  6. complete_account_setup()          — 完成建账（海绵熔断 → ACTIVE）
  7. reopen_account_setup()            — 重新开账（ACTIVE 无凭证 → ONBOARDING）
  8. export_template()                 — 导出 Excel 期初余额模板
  9. import_from_excel()               — 从 Excel 批量导入期初余额

实现的规格点：
  G2.1  1月开账 → ytd 强制归零
  G2.2  year_start_balance 防篡改推导（禁止前端传入）
  G2.3  树状向上自动聚合父科目
  G3.1  四维度本位币试算平衡
  G3.2  外币独立试算平衡
  G4.1  openpyxl 导出标准列 Excel 模板
  G4.2  openpyxl 解析上传 Excel，行号级错误收集
  G5.1  海绵熔断：试算不平则写入 1901，账套置 ACTIVE
  M1    auxiliary_hash MD5 计算规范
  M4    余额方向异常警告（不阻塞保存）
  M5    ACTIVE + 有凭证 → 期初只读；reopen 重新开账
"""
import hashlib
import io
import json
import logging
from datetime import datetime
from typing import Optional

from sqlalchemy.exc import ProgrammingError
from sqlalchemy.orm import Session

from models.account_set import AccountSet, AccountSetStatus
from models.accounting import InitialBalance, SystemSubject, TenantSubject
from schemas.initial_balance_schemas import (
    AuxiliaryEntry,
    BatchSaveInput,
    CompleteAccountSetupResult,
    ForeignTrialBalanceLine,
    InitialBalanceInput,
    InitialBalanceResponse,
    SubjectWithBalance,
    TrialBalanceLine,
    TrialBalanceResult,
)
from services.subject_service import SubjectNotFoundError

logger = logging.getLogger(__name__)


# ── 自定义异常 ──────────────────────────────────────────────────────────────────

class InitialBalanceLockedError(Exception):
    """账套 ACTIVE 且已有凭证，期初余额禁止修改"""
    pass


# ── 工具函数 ────────────────────────────────────────────────────────────────────

def _compute_auxiliary_hash(auxiliary_details: list[AuxiliaryEntry]) -> str:
    """
    M1: 辅助核算特征哈希。
    无辅助（空列表）→ ""
    有辅助 → MD5(sorted JSON of auxiliary_details)
    """
    if not auxiliary_details:
        return ""
    sorted_entries = sorted(
        [{"type": d.type, "id": d.id, "name": d.name} for d in auxiliary_details],
        key=lambda x: (x["type"], x["id"]),
    )
    return hashlib.md5(
        json.dumps(sorted_entries, ensure_ascii=False).encode()
    ).hexdigest()


def _compute_year_start(
    initial: float,
    ytd_debit: float,
    ytd_credit: float,
    direction: str,
) -> float:
    """
    G2.2: 年初余额防篡改公式。
    借方科目：year_start = initial + ytd_credit - ytd_debit
    贷方科目：year_start = initial + ytd_debit  - ytd_credit
    """
    if direction == "借":
        return initial + ytd_credit - ytd_debit
    else:
        return initial + ytd_debit - ytd_credit


def _is_january_start(account_set: AccountSet) -> bool:
    """G2.1: 账套是否 1 月开账（start_period 格式 YYYY-MM，月份 == '01' 为 True）"""
    try:
        month = account_set.start_period.split("-")[1]
        return month == "01"
    except (IndexError, AttributeError):
        return False


def _check_locked(db: Session, tenant_id: int, account_set_id: int) -> None:
    """
    M5: 锁定检查。
    账套 ACTIVE 且存在 voucher_header 记录 → 抛 InitialBalanceLockedError。
    voucher_header 表可能尚未创建（Sprint 3 前），捕获 ProgrammingError 跳过。
    """
    account_set = db.get(AccountSet, account_set_id)
    if account_set is None or account_set.tenant_id != tenant_id:
        raise SubjectNotFoundError(f"账套 {account_set_id} 不存在")

    if account_set.status != AccountSetStatus.ACTIVE:
        return  # ONBOARDING 状态不锁定

    try:
        from sqlalchemy import text
        row = db.execute(
            text(
                "SELECT 1 FROM voucher_header "
                "WHERE tenant_id=:tid AND account_set_id=:asid AND is_deleted=0 "
                "LIMIT 1"
            ),
            {"tid": tenant_id, "asid": account_set_id},
        ).fetchone()
        if row:
            raise InitialBalanceLockedError(
                "账套已激活且存在凭证记录，期初余额禁止修改。如需修改请先执行重新开账。"
            )
    except ProgrammingError:
        # voucher_header 表尚未创建（Sprint 3 前），不锁定
        logger.debug("voucher_header 表不存在，跳过锁定检查")


# ══════════════════════════════════════════════════════════════════════════════
# InitialBalanceService
# ══════════════════════════════════════════════════════════════════════════════

class InitialBalanceService:

    def __init__(self, db: Session) -> None:
        self.db = db

    # ── 内部查询辅助 ───────────────────────────────────────────────────────────

    def _get_account_set(self, tenant_id: int, account_set_id: int) -> AccountSet:
        account_set = self.db.get(AccountSet, account_set_id)
        if account_set is None or account_set.tenant_id != tenant_id:
            raise SubjectNotFoundError(f"账套 {account_set_id} 不存在")
        return account_set

    def _get_subject(
        self, tenant_id: int, account_set_id: int, subject_code: str
    ) -> TenantSubject:
        subject = (
            self.db.query(TenantSubject)
            .filter(
                TenantSubject.tenant_id == tenant_id,
                TenantSubject.account_set_id == account_set_id,
                TenantSubject.subject_code == subject_code,
                TenantSubject.is_deleted == False,
            )
            .first()
        )
        if subject is None:
            raise SubjectNotFoundError(f"科目 {subject_code} 不存在")
        return subject

    def _get_subject_or_none(
        self, tenant_id: int, account_set_id: int, subject_code: str
    ) -> Optional[TenantSubject]:
        return (
            self.db.query(TenantSubject)
            .filter(
                TenantSubject.tenant_id == tenant_id,
                TenantSubject.account_set_id == account_set_id,
                TenantSubject.subject_code == subject_code,
                TenantSubject.is_deleted == False,
            )
            .first()
        )

    # ── 内部写入辅助 ───────────────────────────────────────────────────────────

    def _upsert_balance(
        self,
        tenant_id: int,
        account_set_id: int,
        subject_code: str,
        subject_name: str,
        balance_direction: str,
        initial_balance: float,
        ytd_debit: float,
        ytd_credit: float,
        year_start_balance: float,
        auxiliary_hash: str = "",
        auxiliary_details: Optional[str] = None,
        currency_code: Optional[str] = None,
        foreign_currency_amount: Optional[float] = None,
        exchange_rate: Optional[float] = None,
        quantity: Optional[float] = None,
        unit_price: Optional[float] = None,
        is_ai_sponge: bool = False,
    ) -> InitialBalance:
        """
        按联合唯一键 (tenant_id, account_set_id, subject_code, auxiliary_hash) 做 UPSERT。
        存在则更新，不存在则新建。
        """
        record = (
            self.db.query(InitialBalance)
            .filter(
                InitialBalance.tenant_id == tenant_id,
                InitialBalance.account_set_id == account_set_id,
                InitialBalance.subject_code == subject_code,
                InitialBalance.auxiliary_hash == auxiliary_hash,
            )
            .first()
        )
        if record is None:
            record = InitialBalance(
                tenant_id=tenant_id,
                account_set_id=account_set_id,
                subject_code=subject_code,
                subject_name=subject_name,
                balance_direction=balance_direction,
                auxiliary_hash=auxiliary_hash,
            )
            self.db.add(record)

        record.subject_name = subject_name
        record.balance_direction = balance_direction
        record.initial_balance = initial_balance
        record.ytd_debit = ytd_debit
        record.ytd_credit = ytd_credit
        record.year_start_balance = year_start_balance
        record.auxiliary_details = auxiliary_details
        record.currency_code = currency_code
        record.foreign_currency_amount = foreign_currency_amount
        record.exchange_rate = exchange_rate
        record.quantity = quantity
        record.unit_price = unit_price
        record.is_ai_sponge = is_ai_sponge
        return record

    def _reaggregate_summary(
        self,
        tenant_id: int,
        account_set_id: int,
        subject_code: str,
        subject_name: str,
        balance_direction: str,
    ) -> None:
        """
        重新聚合某科目的汇总记录（auxiliary_hash=""）。
        汇总 = 所有辅助明细记录（auxiliary_hash != ""）的 SUM。
        """
        aux_records = (
            self.db.query(InitialBalance)
            .filter(
                InitialBalance.tenant_id == tenant_id,
                InitialBalance.account_set_id == account_set_id,
                InitialBalance.subject_code == subject_code,
                InitialBalance.auxiliary_hash != "",
            )
            .all()
        )
        total_initial = sum(r.initial_balance for r in aux_records)
        total_ytd_debit = sum(r.ytd_debit for r in aux_records)
        total_ytd_credit = sum(r.ytd_credit for r in aux_records)
        year_start = _compute_year_start(
            total_initial, total_ytd_debit, total_ytd_credit, balance_direction
        )
        self._upsert_balance(
            tenant_id=tenant_id,
            account_set_id=account_set_id,
            subject_code=subject_code,
            subject_name=subject_name,
            balance_direction=balance_direction,
            initial_balance=total_initial,
            ytd_debit=total_ytd_debit,
            ytd_credit=total_ytd_credit,
            year_start_balance=year_start,
            auxiliary_hash="",
        )

    def _propagate_up(
        self,
        tenant_id: int,
        account_set_id: int,
        subject_code: str,
    ) -> None:
        """
        G2.3: 递归向上聚合父科目。
        从 subject_code 的父科目开始，沿 parent_code 链向上，每层重新 SUM 子科目汇总余额。
        """
        current = self._get_subject_or_none(tenant_id, account_set_id, subject_code)
        if current is None or current.parent_code is None:
            return

        parent_code: Optional[str] = current.parent_code
        while parent_code:
            parent_subj = self._get_subject_or_none(tenant_id, account_set_id, parent_code)
            if parent_subj is None:
                break

            # 获取父科目所有直接子科目
            children = (
                self.db.query(TenantSubject)
                .filter(
                    TenantSubject.tenant_id == tenant_id,
                    TenantSubject.account_set_id == account_set_id,
                    TenantSubject.parent_code == parent_code,
                    TenantSubject.is_deleted == False,
                )
                .all()
            )
            child_codes = [c.subject_code for c in children]
            if not child_codes:
                break

            # SUM 所有子科目的汇总记录（auxiliary_hash=""）
            child_balances = (
                self.db.query(InitialBalance)
                .filter(
                    InitialBalance.tenant_id == tenant_id,
                    InitialBalance.account_set_id == account_set_id,
                    InitialBalance.subject_code.in_(child_codes),
                    InitialBalance.auxiliary_hash == "",
                )
                .all()
            )
            total_initial = sum(r.initial_balance for r in child_balances)
            total_ytd_debit = sum(r.ytd_debit for r in child_balances)
            total_ytd_credit = sum(r.ytd_credit for r in child_balances)
            year_start = _compute_year_start(
                total_initial, total_ytd_debit, total_ytd_credit,
                parent_subj.balance_direction,
            )
            self._upsert_balance(
                tenant_id=tenant_id,
                account_set_id=account_set_id,
                subject_code=parent_code,
                subject_name=parent_subj.subject_name,
                balance_direction=parent_subj.balance_direction,
                initial_balance=total_initial,
                ytd_debit=total_ytd_debit,
                ytd_credit=total_ytd_credit,
                year_start_balance=year_start,
                auxiliary_hash="",
            )
            parent_code = parent_subj.parent_code

    # ── 公开方法 ───────────────────────────────────────────────────────────────

    def save_balance(
        self,
        tenant_id: int,
        account_set_id: int,
        inp: InitialBalanceInput,
    ) -> InitialBalanceResponse:
        """
        G2.1 + G2.2 + G2.3 + M1 + M4：保存单条期初余额。

        步骤：
          1. 查科目（404 if not found）
          2. 锁定检查（M5）
          3. 1月开账 → ytd 强制归零（G2.1）
          4. 推导 year_start_balance（G2.2）
          5. 余额方向异常警告（M4，不阻塞）
          6. 数量×单价一致性警告（不阻塞）
          7. 计算 auxiliary_hash（M1）
          8. UPSERT
          9. 若有辅助，重新聚合汇总记录
         10. 递归向上聚合父科目（G2.3）
        """
        # 1. 查科目
        subject = self._get_subject(tenant_id, account_set_id, inp.subject_code)

        # 2. 锁定检查
        _check_locked(self.db, tenant_id, account_set_id)

        # 3. 1月开账 → ytd 归零（G2.1）
        account_set = self._get_account_set(tenant_id, account_set_id)
        ytd_debit = inp.ytd_debit
        ytd_credit = inp.ytd_credit
        if _is_january_start(account_set):
            ytd_debit = 0.0
            ytd_credit = 0.0

        # 4. 推导 year_start_balance（G2.2）
        year_start = _compute_year_start(
            inp.initial_balance, ytd_debit, ytd_credit,
            subject.balance_direction,
        )

        # 5. 余额方向异常警告（M4）
        # year_start < 0 说明该科目出现了反向余额，属于异常
        direction_warning: Optional[str] = None
        if year_start < 0:
            direction_warning = (
                f"科目 {subject.subject_code}（{subject.subject_name}）"
                f"余额方向为【{subject.balance_direction}】，"
                f"但推算出的年初余额为负数（{year_start:.2f}），请检查录入数据是否正确"
            )

        # 6. 数量×单价一致性警告
        if (
            inp.quantity is not None
            and inp.unit_price is not None
            and inp.initial_balance != 0
        ):
            expected = round(inp.quantity * inp.unit_price, 2)
            if abs(expected - inp.initial_balance) > 0.01:
                qty_warning = (
                    f"数量({inp.quantity}) × 单价({inp.unit_price}) = {expected}，"
                    f"与期初余额({inp.initial_balance})不符"
                )
                direction_warning = (
                    f"{direction_warning}；{qty_warning}"
                    if direction_warning
                    else qty_warning
                )

        # 7. 计算 auxiliary_hash（M1）
        aux_hash = _compute_auxiliary_hash(inp.auxiliary_details)
        aux_details_json: Optional[str] = None
        if inp.auxiliary_details:
            aux_details_json = json.dumps(
                [{"type": e.type, "id": e.id, "name": e.name} for e in inp.auxiliary_details],
                ensure_ascii=False,
            )

        # 8. UPSERT
        record = self._upsert_balance(
            tenant_id=tenant_id,
            account_set_id=account_set_id,
            subject_code=subject.subject_code,
            subject_name=subject.subject_name,
            balance_direction=subject.balance_direction,
            initial_balance=inp.initial_balance,
            ytd_debit=ytd_debit,
            ytd_credit=ytd_credit,
            year_start_balance=year_start,
            auxiliary_hash=aux_hash,
            auxiliary_details=aux_details_json,
            currency_code=inp.currency_code,
            foreign_currency_amount=inp.foreign_currency_amount,
            exchange_rate=inp.exchange_rate,
            quantity=inp.quantity,
            unit_price=inp.unit_price,
        )
        self.db.flush()

        # 9. 若有辅助，重新聚合汇总记录（auxiliary_hash=""）
        if aux_hash != "":
            self._reaggregate_summary(
                tenant_id, account_set_id,
                subject.subject_code, subject.subject_name,
                subject.balance_direction,
            )
            self.db.flush()

        # 10. 递归向上聚合父科目（G2.3）
        self._propagate_up(tenant_id, account_set_id, subject.subject_code)
        self.db.flush()

        # 构造返回
        aux_list: list[AuxiliaryEntry] = (
            [AuxiliaryEntry(**e) for e in json.loads(aux_details_json)]
            if aux_details_json
            else []
        )
        return InitialBalanceResponse(
            id=record.id,
            subject_code=record.subject_code,
            subject_name=record.subject_name,
            balance_direction=record.balance_direction,
            initial_balance=record.initial_balance,
            ytd_debit=record.ytd_debit,
            ytd_credit=record.ytd_credit,
            year_start_balance=record.year_start_balance,
            currency_code=record.currency_code,
            foreign_currency_amount=record.foreign_currency_amount,
            exchange_rate=record.exchange_rate,
            quantity=record.quantity,
            unit_price=record.unit_price,
            auxiliary_details=aux_list,
            auxiliary_hash=record.auxiliary_hash,
            is_ai_sponge=record.is_ai_sponge,
            direction_warning=direction_warning,
        )

    def batch_save(
        self,
        tenant_id: int,
        account_set_id: int,
        inp: BatchSaveInput,
    ) -> dict:
        """
        M2: 批量保存，整体事务。
        一行失败 → 全部回滚。
        """
        saved = 0
        warnings: list[str] = []
        try:
            for row in inp.rows:
                resp = self.save_balance(tenant_id, account_set_id, row)
                saved += 1
                if resp.direction_warning:
                    warnings.append(f"{row.subject_code}: {resp.direction_warning}")
            self.db.commit()
        except Exception:
            self.db.rollback()
            raise
        return {"saved": saved, "warnings": warnings}

    def get_balances_with_subjects(
        self,
        tenant_id: int,
        account_set_id: int,
    ) -> list[SubjectWithBalance]:
        """
        M3: 科目树 + 期初余额联合查询。
        一次 SELECT TenantSubject + 一次 SELECT InitialBalance，Python 内存构建树。
        只取 auxiliary_hash="" 的汇总记录。
        """
        subjects = (
            self.db.query(TenantSubject)
            .filter(
                TenantSubject.tenant_id == tenant_id,
                TenantSubject.account_set_id == account_set_id,
                TenantSubject.is_deleted == False,
            )
            .order_by(TenantSubject.sort_order, TenantSubject.subject_code)
            .all()
        )
        balances = (
            self.db.query(InitialBalance)
            .filter(
                InitialBalance.tenant_id == tenant_id,
                InitialBalance.account_set_id == account_set_id,
                InitialBalance.auxiliary_hash == "",
            )
            .all()
        )
        balance_map: dict[str, InitialBalance] = {b.subject_code: b for b in balances}
        parent_codes = {s.parent_code for s in subjects if s.parent_code}

        nodes: dict[str, SubjectWithBalance] = {}
        for s in subjects:
            b = balance_map.get(s.subject_code)
            nodes[s.subject_code] = SubjectWithBalance(
                subject_code=s.subject_code,
                subject_name=s.subject_name,
                category=s.category,
                balance_direction=s.balance_direction,
                level=s.level,
                has_children=s.subject_code in parent_codes,
                initial_balance=b.initial_balance if b else 0.0,
                ytd_debit=b.ytd_debit if b else 0.0,
                ytd_credit=b.ytd_credit if b else 0.0,
                year_start_balance=b.year_start_balance if b else 0.0,
                is_ai_sponge=b.is_ai_sponge if b else False,
            )

        roots: list[SubjectWithBalance] = []
        for s in subjects:
            node = nodes[s.subject_code]
            if s.parent_code and s.parent_code in nodes:
                nodes[s.parent_code].children.append(node)
            elif s.parent_code is None:
                roots.append(node)
        return roots

    def calculate_trial_balance(
        self,
        tenant_id: int,
        account_set_id: int,
    ) -> TrialBalanceResult:
        """
        G3.1: 本位币四维度试算平衡。
        只取 level=1 科目的汇总记录（auxiliary_hash=""）。
        四维度：期初余额 / 本年累计借方 / 本年累计贷方 / 年初余额。
        每维度：借方余额方向科目合计 vs 贷方余额方向科目合计，差额应为 0。
        """
        level1_subjects = (
            self.db.query(TenantSubject)
            .filter(
                TenantSubject.tenant_id == tenant_id,
                TenantSubject.account_set_id == account_set_id,
                TenantSubject.level == 1,
                TenantSubject.is_deleted == False,
            )
            .all()
        )
        codes = [s.subject_code for s in level1_subjects]
        direction_map = {s.subject_code: s.balance_direction for s in level1_subjects}

        balances = (
            self.db.query(InitialBalance)
            .filter(
                InitialBalance.tenant_id == tenant_id,
                InitialBalance.account_set_id == account_set_id,
                InitialBalance.subject_code.in_(codes),
                InitialBalance.auxiliary_hash == "",
            )
            .all()
        )
        bmap: dict[str, InitialBalance] = {b.subject_code: b for b in balances}

        def _sum(field: str, direction: str) -> float:
            return sum(
                getattr(bmap[c], field)
                for c in codes
                if c in bmap and direction_map[c] == direction
            )

        dims = [
            ("期初余额",    "initial_balance",    "initial_balance"),
            ("本年累计借方", "ytd_debit",          "ytd_debit"),
            ("本年累计贷方", "ytd_credit",         "ytd_credit"),
            ("年初余额",    "year_start_balance", "year_start_balance"),
        ]
        lines: list[TrialBalanceLine] = []
        for label, debit_field, credit_field in dims:
            td = _sum(debit_field, "借")
            tc = _sum(credit_field, "贷")
            diff = round(td - tc, 2)
            lines.append(TrialBalanceLine(
                dimension=label,
                total_debit=td,
                total_credit=tc,
                difference=diff,
                is_balanced=(diff == 0),
            ))

        is_balanced = all(l.is_balanced for l in lines)
        sponge_amount = max(abs(l.difference) for l in lines) if lines else 0.0
        return TrialBalanceResult(
            lines=lines,
            is_balanced=is_balanced,
            sponge_amount=sponge_amount,
        )

    def calculate_foreign_trial_balance(
        self,
        tenant_id: int,
        account_set_id: int,
        currency_code: str,
    ) -> ForeignTrialBalanceLine:
        """
        G3.2: 外币独立试算平衡。
        按 currency_code 筛选，对 foreign_currency_amount 按余额方向求借贷合计。
        """
        records = (
            self.db.query(InitialBalance)
            .filter(
                InitialBalance.tenant_id == tenant_id,
                InitialBalance.account_set_id == account_set_id,
                InitialBalance.currency_code == currency_code,
                InitialBalance.auxiliary_hash == "",
            )
            .all()
        )
        total_debit = sum(
            (r.foreign_currency_amount or 0.0) for r in records
            if r.balance_direction == "借"
        )
        total_credit = sum(
            (r.foreign_currency_amount or 0.0) for r in records
            if r.balance_direction == "贷"
        )
        difference = round(total_debit - total_credit, 2)
        return ForeignTrialBalanceLine(
            currency_code=currency_code,
            total_debit=total_debit,
            total_credit=total_credit,
            difference=difference,
            is_balanced=(difference == 0),
        )

    def complete_account_setup(
        self,
        tenant_id: int,
        account_set_id: int,
    ) -> CompleteAccountSetupResult:
        """
        G5.1: 完成建账海绵熔断。
        1. 试算平衡
        2. 不平衡 → 写入 1901 配平（is_ai_sponge=True）
        3. 账套置 ACTIVE，activated_at = now()
        """
        account_set = self._get_account_set(tenant_id, account_set_id)
        trial = self.calculate_trial_balance(tenant_id, account_set_id)

        sponge_amount = 0.0
        sponge_subject: Optional[str] = None
        was_balanced = trial.is_balanced

        if not trial.is_balanced:
            # 取期初余额维度的差额：total_debit - total_credit
            initial_line = next(l for l in trial.lines if l.dimension == "期初余额")
            sponge = initial_line.difference  # 正 → 借方多；负 → 贷方多

            # 确保 1901 科目存在于 TenantSubject
            subj_1901 = self._get_subject_or_none(tenant_id, account_set_id, "1901")
            if subj_1901 is None:
                sys_1901 = (
                    self.db.query(SystemSubject)
                    .filter(SystemSubject.subject_code == "1901")
                    .first()
                )
                if sys_1901 is None:
                    raise ValueError(
                        "系统科目库中未找到 1901（待处理财产损溢），请先执行科目库初始化"
                    )
                subj_1901 = TenantSubject(
                    tenant_id=tenant_id,
                    account_set_id=account_set_id,
                    subject_code=sys_1901.subject_code,
                    subject_name=sys_1901.subject_name,
                    parent_code=sys_1901.parent_code,
                    category=sys_1901.category,
                    balance_direction=sys_1901.balance_direction,
                    level=sys_1901.level,
                    sort_order=sys_1901.sort_order,
                    system_subject_code=sys_1901.subject_code,
                )
                self.db.add(subj_1901)
                self.db.flush()

            # 配平方向：
            #   sponge > 0（借方多） → 1901 记录挂贷方，填平贷方缺口
            #   sponge < 0（贷方多） → 1901 记录挂借方，填平借方缺口
            abs_sponge = abs(sponge)
            fill_direction = "贷" if sponge > 0 else "借"

            sponge_record = (
                self.db.query(InitialBalance)
                .filter(
                    InitialBalance.tenant_id == tenant_id,
                    InitialBalance.account_set_id == account_set_id,
                    InitialBalance.subject_code == "1901",
                    InitialBalance.auxiliary_hash == "",
                )
                .first()
            )
            if sponge_record is None:
                sponge_record = InitialBalance(
                    tenant_id=tenant_id,
                    account_set_id=account_set_id,
                    subject_code="1901",
                    subject_name=subj_1901.subject_name,
                    auxiliary_hash="",
                )
                self.db.add(sponge_record)

            sponge_record.balance_direction = fill_direction
            sponge_record.initial_balance = abs_sponge
            sponge_record.ytd_debit = 0.0
            sponge_record.ytd_credit = 0.0
            sponge_record.year_start_balance = abs_sponge
            sponge_record.is_ai_sponge = True
            self.db.flush()

            # 向上聚合 1901 的父科目
            self._propagate_up(tenant_id, account_set_id, "1901")
            self.db.flush()

            sponge_amount = abs_sponge
            sponge_subject = "1901"

        # 账套置 ACTIVE
        account_set.status = AccountSetStatus.ACTIVE
        account_set.activated_at = datetime.utcnow()
        self.db.commit()

        return CompleteAccountSetupResult(
            success=True,
            account_set_id=account_set_id,
            final_status=AccountSetStatus.ACTIVE,
            was_balanced=was_balanced,
            sponge_amount=sponge_amount,
            sponge_subject=sponge_subject,
            message=(
                "建账完成，试算平衡，账套已激活"
                if was_balanced
                else f"试算不平衡，已自动写入 1901 配平（差额 {sponge_amount:.2f}），账套已激活"
            ),
        )

    def reopen_account_setup(
        self,
        tenant_id: int,
        account_set_id: int,
    ) -> dict:
        """
        M5: 重新开账。
        仅 ACTIVE 且无凭证时允许，账套置回 ONBOARDING。
        """
        account_set = self._get_account_set(tenant_id, account_set_id)

        if account_set.status != AccountSetStatus.ACTIVE:
            raise ValueError(
                f"账套当前状态为 {account_set.status}，只有 ACTIVE 状态才能重新开账"
            )

        # 检查是否有凭证（有则拒绝，HTTP 409 由路由层处理）
        try:
            from sqlalchemy import text
            row = self.db.execute(
                text(
                    "SELECT 1 FROM voucher_header "
                    "WHERE tenant_id=:tid AND account_set_id=:asid AND is_deleted=0 "
                    "LIMIT 1"
                ),
                {"tid": tenant_id, "asid": account_set_id},
            ).fetchone()
            if row:
                raise InitialBalanceLockedError(
                    "账套已存在凭证记录，禁止重新开账。如需调整请先删除相关凭证。"
                )
        except ProgrammingError:
            pass  # 表不存在，允许继续

        account_set.status = AccountSetStatus.ONBOARDING
        account_set.activated_at = None
        self.db.commit()
        return {"success": True, "message": "账套已重置为建账状态，可重新编辑期初余额"}

    def export_template(
        self,
        tenant_id: int,
        account_set_id: int,
    ) -> bytes:
        """
        G4.1: 导出期初余额 Excel 模板。
        列：科目编码 | 科目名称 | 余额方向 | 期初余额 | 本年累计借方 | 本年累计贷方
        父科目行灰色背景 + 提示文字；叶子科目行空白待填。
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl 未安装，无法导出 Excel 模板")

        subjects = (
            self.db.query(TenantSubject)
            .filter(
                TenantSubject.tenant_id == tenant_id,
                TenantSubject.account_set_id == account_set_id,
                TenantSubject.is_deleted == False,
            )
            .order_by(TenantSubject.sort_order, TenantSubject.subject_code)
            .all()
        )
        parent_codes = {s.parent_code for s in subjects if s.parent_code}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "期初余额"

        headers = ["科目编码", "科目名称", "余额方向", "期初余额", "本年累计借方", "本年累计贷方"]
        ws.append(headers)

        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        gray_font = Font(color="808080", italic=True)

        for s in subjects:
            is_parent = s.subject_code in parent_codes
            if is_parent:
                ws.append([s.subject_code, s.subject_name, s.balance_direction,
                            "汇总行，请勿手工填写", "", ""])
                for cell in ws[ws.max_row]:
                    cell.fill = gray_fill
                    cell.font = gray_font
            else:
                ws.append([s.subject_code, s.subject_name, s.balance_direction,
                            "", "", ""])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def import_from_excel(
        self,
        tenant_id: int,
        account_set_id: int,
        file_bytes: bytes,
    ) -> dict:
        """
        G4.2: 从 Excel 导入期初余额。
        按「科目编码」列匹配 TenantSubject，找不到编码收集错误不中断其他行。
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl 未安装，无法解析 Excel 文件")

        wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        # 解析表头
        headers = [
            (str(cell.value).strip() if cell.value else "")
            for cell in ws[1]
        ]
        col_map: dict[str, int] = {h: i for i, h in enumerate(headers) if h}

        required = ["科目编码", "期初余额", "本年累计借方", "本年累计贷方"]
        missing = [c for c in required if c not in col_map]
        if missing:
            return {
                "imported": 0,
                "errors": [{"row": 0, "reason": f"缺少必要列：{missing}"}],
                "warnings": [],
            }

        imported = 0
        errors: list[dict] = []
        warnings: list[str] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_code = row[col_map["科目编码"]]
            if not raw_code:
                continue
            subject_code = str(raw_code).strip()

            # 跳过汇总行提示（灰色行）
            raw_initial = row[col_map["期初余额"]]
            if isinstance(raw_initial, str) and "请勿" in raw_initial:
                continue

            try:
                initial_balance = float(raw_initial or 0)
                ytd_debit = float(row[col_map["本年累计借方"]] or 0)
                ytd_credit = float(row[col_map["本年累计贷方"]] or 0)
            except (ValueError, TypeError) as exc:
                errors.append({"row": row_idx, "reason": f"数值格式错误：{exc}"})
                continue

            inp = InitialBalanceInput(
                subject_code=subject_code,
                initial_balance=initial_balance,
                ytd_debit=ytd_debit,
                ytd_credit=ytd_credit,
            )
            try:
                resp = self.save_balance(tenant_id, account_set_id, inp)
                imported += 1
                if resp.direction_warning:
                    warnings.append(f"第{row_idx}行 {subject_code}: {resp.direction_warning}")
            except SubjectNotFoundError as exc:
                errors.append({"row": row_idx, "reason": f"科目不存在：{exc}"})
            except Exception as exc:
                errors.append({"row": row_idx, "reason": str(exc)})

        if imported > 0:
            self.db.flush()

        return {"imported": imported, "errors": errors, "warnings": warnings}
